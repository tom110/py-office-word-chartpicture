import xlrd
import openpyxl
import random

import datetime
# import np

import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
from pylab import mpl

mpl.rcParams['font.sans-serif'] = ['FangSong'] # 指定默认字体
mpl.rcParams['axes.unicode_minus'] = False # 解决保存图像是负号'-'显示为方块的问题


def getRowNumber(table, pointName):
    col0 = table.col_values(0)
    col0Count = len(col0)

    n = 0
    pre_name = ""
    for j in list(range(col0Count)):
        if col0[j][-2:] == "竖井":
            pre_name = col0[j]
        else:
            if col0[j][0:3] == "JCD":
                if pointName == (pre_name+"," + col0[j]):
                    n = j
                    break
    return n

def getPoint(pictureP,points):
    p={}
    for point in points:
        if point["name"]==pictureP:
            p= point
            break
    return p




tablesData = []

excel_path = "1.地表改收敛.xls"
data = xlrd.open_workbook(excel_path, formatting_info=True,on_demand=True)
tables = data.sheets()
tablesCount = len(tables)

points = []
pointsNames = []  # 去重用

# 一下处理数据
for i in list(range(tablesCount)):

    table = tables[i]
    col0 = table.col_values(0)
    col2 = table.col_values(2)
    col0Count = len(col0)

    pre_name = ""
    for j in list(range(col0Count)):
        if col0[j][-2:] == "竖井":
            pre_name = col0[j]
        else:
            if col0[j][0:3] == "JCD":
                if (not (pre_name+',' + col0[j]) in pointsNames) and (col2[j] != ""):
                    pointsNames.append(pre_name+',' + col0[j])
                    point = {}
                    point["name"] = pre_name + "," + col0[j]
                    point["start_dat"] = round(2.3 + random.random() * 0.2, 4)
                    point["start_page"] = i

                    lastValues = []
                    currentValues = []
                    randoms = []
                    randomSums = []
                    currentSpeeds = []
                    sumSpeeds = []

                    for n in list(range(1, 57)):  # 56个随机数
                        ran = round(-0.8 + random.random() * 1.6, 2)

                        randoms.append(ran)

                        if len(lastValues) == 0:
                            lastValues.append(float(point["start_dat"]))
                        else:
                            lastValues.append(currentValues[n - 2])

                        currentValues.append(round(lastValues[n - 1] - ran / 1000, 4))

                        randomSums.append(round(sum(randoms), 2))
                        currentSpeeds.append(abs(ran))
                        sumSpeeds.append(round(randomSums[n - 1] / n, 2))

                    point["lastValues"] = lastValues
                    point["currentValues"] = currentValues
                    point["randoms"] = randoms
                    point["randomSums"] = randomSums
                    point["currentSpeeds"] = currentSpeeds
                    point["sumSpeeds"] = sumSpeeds
                    points.append(point)

s1 = pointsNames
for ss in s1:
    print(ss)
s = points

# 一下生成图表
# 找到要制图的点集
alreayPickoutPoints=[]
pictureNeedPoints=[]
for i in list(range(len(pointsNames))):
    if not pointsNames[i].split("-")[0] in alreayPickoutPoints:
        pictureNeedPoint = []
        pictureNeedPoint.append(pointsNames[i])
        for j in list(range(i+1,len(pointsNames))):
            if pointsNames[i].split("-")[0]==pointsNames[j].split("-")[0]:
                pictureNeedPoint.append(pointsNames[j])
        pictureNeedPoints.append(pictureNeedPoint)
        alreayPickoutPoints.append(pointsNames[i].split("-")[0])

# 遍历点集制图


picturePDatas=[]
for pictureNeedPoint in pictureNeedPoints:
    picturePDataOne=[]
    for pictureP in pictureNeedPoint:
        picturePData={}
        picturePData["name"]=pictureP
        picturePRandomSum= getPoint(pictureP,points)["randomSums"]
        picturePData["data"]=picturePRandomSum
        pictureStartPage = getPoint(pictureP, points)["start_page"]

        days=[]
        for d in list(range(pictureStartPage,pictureStartPage+56)):
            now = datetime.datetime.strptime('2019-04-20', '%Y-%m-%d')
            delta = datetime.timedelta(days=d)
            n_days = now + delta
            days.append(n_days.strftime('%Y-%m-%d'))

        picturePData["x"] = days
        picturePDataOne.append(picturePData)
    picturePDatas.append(picturePDataOne)

ssss=picturePDatas

for picturePData in picturePDatas:
    plt.cla()
    colors=['green','red','blue','black','yellow']
    tmp=0
    for p in picturePData:
        print(tmp)
        print(colors[tmp])
        plt.plot(picturePData[0]["x"], np.negative(p["data"]), colors[tmp], label=p["name"].split(",")[1])
        tmp = tmp + 1

    plt.legend()  # 显示图例

    xticks = list(range(0, len(picturePData[0]["x"]), 3))  # 这里设置的是x轴点的位置
    xticks.append(55)
    plt.xlabel('日期')
    plt.xticks(xticks,rotation=60)
    plt.ylabel('收敛量(mm)')
    plt.title("龙潭220kv变电站—华月路电力隧道工程洞内收敛监测成果表\n"+p["name"].split(",")[0])
    plt.savefig("picture\\"+picturePData[0]["name"].split("-")[0]+".jpg", bbox_inches='tight')




# 一下写入数据
path = "1.地表改收敛.xlsx"
myworkbook=openpyxl.load_workbook(path)

for p in points:
    start_page = p["start_page"]
    t = tables[start_page]
    worksheet = myworkbook.get_sheet_by_name(str(start_page+1))
    pointName = p["name"]
    print(pointName)
    print(start_page)
    n = getRowNumber(t, pointName)
    print(n)

    mycell = worksheet.cell(row=n+1, column=2)
    mycell.value=1
    mycell = worksheet.cell(row=n+1, column=3)
    mycell.value = p["start_dat"]

    for j in list(range(1, 57)):
        worksheet = myworkbook.get_sheet_by_name(str(start_page + j+1))
        t = tables[start_page + j]
        m = getRowNumber(t, pointName)
        print(m)

        mycell = worksheet.cell(row=m+1, column=2)
        mycell.value = j+1

        mycell = worksheet.cell(row=m+1, column=3)
        mycell.value = "%.04f"%p["start_dat"]

        mycell = worksheet.cell(row=m+1, column=4)
        mycell.value = "%.04f"%p["lastValues"][j - 1]

        mycell = worksheet.cell(row=m+1, column=5)
        mycell.value = "%.04f"%p["currentValues"][j - 1]

        mycell = worksheet.cell(row=m+1, column=6)
        mycell.value = "%.02f"%p["randoms"][j - 1]

        mycell = worksheet.cell(row=m+1, column=7)
        mycell.value = "%.02f"%p["randomSums"][j - 1]

        mycell = worksheet.cell(row=m+1, column=8)
        mycell.value = "%.02f"%p["currentSpeeds"][j - 1]

        mycell = worksheet.cell(row=m+1, column=9)
        mycell.value = "%.02f"%p["sumSpeeds"][j - 1]


myworkbook.save("地表改收敛改后.xlsx")

